"""
views.py

This script handles user interactions and data processing in a Django web application, primarily focusing on
session management and analysis execution using the OpenAI API.

Main Elements:
- Session Management: Manages the storage and retrieval of analysis-related data in user sessions.
- Analysis Execution: Initiates and manages asynchronous analysis tasks, including the creation of flowcharts.

Additional Functionalities:
- Clears session data to maintain data integrity.
- Generates Excel reports based on analysis results.
- Supports asynchronous communication with WebSocket channels.
- Handles user authentication, including login and logout processes.
- Manages file uploads and saves them to the server.
- Parses and processes JSON data to generate tables and flowcharts.
- Saves generated flowcharts as images.
- Updates session data via POST requests to ensure consistency across user interactions.
"""

from asgiref.sync import async_to_sync
from botocore.exceptions import ClientError
from channels.layers import get_channel_layer
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.core.serializers.json import DjangoJSONEncoder
from graphviz import Digraph
from openpyxl import Workbook
from urllib.parse import quote
from qda_gpt.prompts.prompts_ca import ca_instruction
from qda_gpt.prompts.prompts_gt import gt_instruction
from qda_gpt.prompts.prompts_ta import ta_instruction
from qda_gpt.analyses import thematic_analysis, content_analysis, grounded_theory
from .openai_api import initialize_openai_resources, create_thread
from .forms import LoginForm, SetupForm
from .__version__ import __version__
import asyncio
import pandas as pd
import inspect
import os
import time
import json
import logging
import boto3



logger = logging.getLogger(__name__)

# Handle the user login process.
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        # Authenticate the user with the provided credentials
        user = authenticate(request, username=username, password=password)
        if user is not None:
            # Log in the user and redirect to the dashboard
            login(request, user)
            return redirect('dashboard')
        else:
            return render(request, 'registration/login.html', {'error': 'Invalid credentials'})
    return render(request, 'registration/login.html')

# Handle the user logout process.
def logout_view(request):
    logout(request)
    return render(request, 'registration/logout.html')

# Clear specific session data related to the analysis.
def clear_session_data(request):
    session_keys = [
        'response', 'deletion_results', 'console_output', 'analysis_status', 'uploaded_file_name', 'first_response',
        'second_response', 'third_response', 'fourth_response', 'fifth_response', 'sixth_response', 'seventh_response',
        'eighth_response', 'analysis_type', 'user_prompt', 'file_name', 'tables', 'prompt_table_pairs', 'flowchart_path'
    ]
    # Remove each key from the session if it exists
    for key in session_keys:
        request.session.pop(key, None)
    request.session.save()

# Clear the session data and redirect to the dashboard.
def clear_session(request):
    clear_session_data(request)
    return redirect('dashboard')

# Retrieve the current status of the analysis from the session.
def get_analysis_status(request):
    status = request.session.get('analysis_status', '')
    logger.debug(f"Current analysis_status: {status}\n")
    return JsonResponse({'analysis_status': status})

# Handle file upload and save it to the server.
def handle_uploaded_file(f):
    file_path = os.path.join('uploads', f.name)
    # Create the 'uploads' directory if it doesn't exist
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    # Save the uploaded file to the specified path
    with open(file_path, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    return file_path

# Generate tables from JSON response text.
def generate_tables_from_response(response_text):
    try:
        # Extract JSON content from the response text
        start = response_text.find('{')
        end = response_text.rfind('}') + 1
        if start == -1 or end == -1:
            raise json.JSONDecodeError("Invalid JSON format", response_text, 0)
        response_text = response_text[start:end]
        response_json = json.loads(response_text)

        tables = []

        # Process the JSON content into tables
        if isinstance(response_json, dict):
            for table_name, records in response_json.items():
                if isinstance(records, list):
                    # Normalize the JSON data, flattening nested structures with a custom separator
                    df = pd.json_normalize(records, sep='_')
                    # Explode any nested columns to further flatten the data
                    df = explode_nested_columns(df)
                    # Append the processed table data to the 'tables' list
                    tables.append({
                        'table_name': table_name,
                        'columns': df.columns.tolist(),
                        'data': df.values.tolist()
                    })
                # If the value is a dictionary (indicating a single record for a table)
                elif isinstance(records, dict):
                    # Normalize the JSON data
                    df = pd.json_normalize(records, sep='_')
                    # Explode any nested columns
                    df = explode_nested_columns(df)
                    if df.shape[0] == 1:
                        df = df.T.reset_index()
                        df.columns = ['Field', 'Value']
                    # Append the processed table data to the 'tables' list
                    tables.append({
                        'table_name': table_name,
                        'columns': df.columns.tolist(),
                        'data': df.values.tolist()
                    })
        return tables

    except json.JSONDecodeError as e:
        logger.error(f"JSONDecodeError: {e}\n")
        return []
    except Exception as e:
        logger.error(f"An error occurred: {e}\n")
        return []

# Explode and normalize nested columns in the DataFrame.
def explode_nested_columns(df):
    """
    Explode and normalize nested columns in the DataFrame.
    """
    for col in df.columns:
        # Handle lists within columns by exploding them into separate rows
        if isinstance(df[col].iloc[0], list):
            df = df.explode(col).reset_index(drop=True)
        # Handle dictionaries within columns by normalizing them into separate columns
        if isinstance(df[col].iloc[0], dict):
            df = df.drop(columns=[col]).join(df[col].apply(pd.Series).add_prefix(f"{col}_"))
    return df

# Generate and download an Excel file based on the analysis results.
def download_xlsx(request):
    analysis_type = request.session.get('analysis_type', 'N/A')
    user_prompt = request.session.get('user_prompt', 'N/A')
    file_name = request.GET.get('file_name', 'qda.xlsx')  # Get the filename from the query parameters

    # Map analysis type to full names
    analysis_type_full_name = {
        'thematic': 'Thematic Analysis',
        'grounded': 'Grounded Theory',
        'content': 'Content Analysis'
    }.get(analysis_type, 'Unknown Analysis Type')

    # Select the correct instructions based on the analysis type
    instructions_template = {
        'Thematic Analysis': ta_instruction,
        'Grounded Theory': gt_instruction,
        'Content Analysis': ca_instruction
    }.get(analysis_type_full_name, 'N/A')

    instructions = instructions_template.format(user_prompt=user_prompt)

    logger.debug(f"[DEBUG] Session prompt_table_pairs: {request.session.get('prompt_table_pairs', [])}")
    prompt_table_pairs = request.session.get('prompt_table_pairs', [])

    # Sanitize the filename for use in Content-Disposition
    file_name = quote(file_name)

    # Create the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"

    # Write analysis type and instructions in the first sheet
    ws.append(['Analysis Type', analysis_type_full_name])
    ws.append([])  # Add an empty row for separation
    ws.append(['Instructions'])
    for line in instructions.split('\n'):
        ws.append([line])
    ws.append([])  # Add an empty row for separation

    # Add each prompt and its associated response to separate sheets
    for index, pair in enumerate(prompt_table_pairs):
        # Create a new sheet for the prompt
        prompt_sheet_title = f'Prompt {index + 1}'
        prompt_ws = wb.create_sheet(title=prompt_sheet_title)  # Create the sheet in the workbook
        prompt_ws.append(['Prompt'])  # Add a header for the prompt content
        # Add each line of the prompt text to the sheet
        for line in pair.get('prompt', 'N/A').split('\n'):
            prompt_ws.append([line])

        # Create a new sheet for the response
        response_sheet_title = f'Response {index + 1}'
        response_ws = wb.create_sheet(title=response_sheet_title)  # Create the sheet in the workbook
        # Iterate through each table in the response
        for table in pair['tables']:
            response_ws.append([table['table_name']])  # Add the table name as a header
            response_ws.append(table['columns'])  # Add the column names as the second row
            # Add each row of data from the table
            for row in table['data']:
                if isinstance(row, list):
                    response_ws.append(row)
                elif isinstance(row, dict):
                    response_ws.append([row.get(col, '') for col in table['columns']])
            response_ws.append([])  # Add an empty row for separation

    wb.save(response)

    return response

# Wrap text to a specified maximum length.
def wrap_text(text, max_length):
    words = text.split()
    lines = []
    current_line = ""
    for word in words:
        # Add words to the current line until the max length is reached
        if len(current_line) + len(word) + 1 <= max_length:
            if current_line:
                current_line += " "
            current_line += word
        else:
            lines.append(current_line)
            current_line = word
    if current_line:
        lines.append(current_line)
    return "\n".join(lines)

# Create a combined flowchart from the provided data using Graphviz.
def create_combined_flowchart(data):
    logger.debug(f"[DEBUG] create_combined_flowchart received data: {data}\n")  # Debug print

    # Clean and parse JSON data
    start = data.find('{')
    end = data.rfind('}') + 1
    if start == -1 or end == -1:
        raise json.JSONDecodeError("Invalid JSON format", data, 0)
    data = data[start:end]
    json_data = json.loads(data)

    # Filter the data to only include tables with "visualization" in their name
    filtered_data = {k: v for k, v in json_data.items() if "visualization" in k}
    logger.debug(f"[DEBUG] Filtered data: {filtered_data}\n")  # Debug print

    # Create a single Digraph instance
    dot = Digraph()

    # Set global graph attributes for spacing
    dot.attr(rankdir='TB')  # Ensure top-to-bottom direction for entire graph

    # Process each CoreCategory in the filtered tables
    for i, table_data in enumerate(filtered_data["table_format_visualization"]):
        core_category = table_data["core_category"]
        relationships = table_data["relationships"]

        # Add a subgraph for each CoreCategory to maintain separation
        with dot.subgraph(name=f'cluster_{i}') as sub:
            sub.attr(label=core_category, rank='same', style='invis')
            nodes = set()
            for relation in relationships:
                description = wrap_text(relation["Description"], 40)
                # Create a left-aligned label with HTML-like line breaks
                formatted_description = '<' + description.replace('\n', '<br align="left"/>') + '>'
                sub.edge(relation["From"], relation["To"], label=formatted_description)
                nodes.add(relation["From"])
                nodes.add(relation["To"])

            # Set all nodes to be rectangles
            for node in nodes:
                sub.node(node, shape='rect')

    return dot


# Save the generated flowchart as a PNG file and optionally upload it to S3.
def save_flowchart_as_png(dot):
    try:
        # Ensure the 'flowcharts' directory exists under 'media'
        flowcharts_dir = os.path.join(settings.MEDIA_ROOT, 'flowcharts')
        os.makedirs(flowcharts_dir, exist_ok=True)

        # Construct the correct full filename (avoid double nesting)
        full_filename = os.path.join(flowcharts_dir, 'flowchart')

        # Render the flowchart and save it as a PNG file locally
        dot.render(full_filename, format='png', cleanup=True)
        logger.debug(f"[DEBUG] Combined flowchart image generated and saved as {full_filename}.png")

        # If using S3, upload the file to the S3 bucket
        if 'DYNO' in os.environ:
            logger.debug(f"Heroku environment recognized. Initializing upload to S3")
            s3_client = boto3.client('s3')  # Initialize an S3 client using Boto3
            try:
                # Attempt to upload the PNG file to the specified S3 bucket
                s3_client.upload_file(f'{full_filename}.png', settings.AWS_STORAGE_BUCKET_NAME,
                                      f'media/flowcharts/flowchart.png')
                logger.debug(f"Flowchart successfully uploaded to S3: media/flowcharts/flowchart.png")
            except ClientError as e:
                logger.error(f"Failed to upload flowchart to S3: {e}")
    except Exception as e:
        logger.error(f"[DEBUG] Error saving flowchart as PNG: {e}")

# Handle the setup process for analysis, including file upload and resource initialization.
def handle_setup(request, setup_form):
    if setup_form.is_valid():
        file = request.FILES.get('file')
        if not file:
            return "Please select a file."
        model_choice = setup_form.cleaned_data['model_choice']
        user_prompt = request.POST.get('user_prompt', '')
        file_path = handle_uploaded_file(file)
        try:
            # Create thread and assign correct thread ID
            request.session['analysis_status'] = "Initializing OpenAI Assistant."
            request.session.save()  # Explicitly save the session
            time.sleep(0.25)
            thread_id = create_thread()
            if thread_id:
                request.session['thread_id'] = thread_id
                resources = initialize_openai_resources(
                    file_path, model_choice, request.session['analysis_type'], user_prompt
                )
                request.session['analysis_status'] = "OpenAI Assistant initialized successfully. Running analysis. This will take a while."
                request.session.save()  # Explicitly save the session

                time.sleep(1)

                print("Waiting for indexing: ", end='', flush=True)
                for i in range(5, -1, -1):
                    print(f"{i} ", end='', flush=True)
                    time.sleep(1)
                    print('\rWaiting for indexing: ', end='',
                          flush=True)  # Return to the beginning of the line and overwrite
                print("0")
                print("Indexing complete.\n")

                # Update session with initialization results
                request.session['initialized'] = True
                request.session['vector_store_id'] = resources['vector_store'].id
                request.session['assistant_id'] = resources['assistant'].id
                request.session['file_id'] = resources['file'].id
                request.session['file_name'] = file.name
                request.session['model_choice'] = model_choice
                request.session['user_prompt'] = user_prompt

                return True  # Indicate success immediately
            else:
                request.session['analysis_status'] = "Failed to create thread."
                request.session.save()  # Explicitly save the session
                return False # Indicate failure immediately

        except Exception as e:
            request.session['analysis_status'] = f"Error initializing OpenAI resources: {str(e)}"
            request.session.save()  # Explicitly save the session
            return False
    return False

# Asynchronously run the selected analysis based on the analysis type.
async def run_analysis_async(analysis_data):
    analysis_type = analysis_data['analysis_type']
    analysis_funcs = {
        'thematic': thematic_analysis,
        'content': content_analysis,
        'grounded': grounded_theory
    }

    # Get the channel layer for handling WebSocket connections or asynchronous communication
    channel_layer = get_channel_layer()

    if analysis_type in analysis_funcs:
        analysis_module = analysis_funcs[analysis_type]
        formatted_prompts = []
        prompt_table_pairs = []
        phases = [func for name, func in analysis_module.__dict__.items() if name.startswith('phase')]
        responses = {}
        flowchart_path = None

        for idx, phase in enumerate(phases):
            # Add delay before each phase starts
            await asyncio.sleep(0.5)

            # Send phase update through WebSocket
            await channel_layer.group_send(
                "analysis_group",
                {
                    "type": "send_analysis_result",
                    "content": {"analysis_status": f"Performing phase {idx + 1} of the analysis..."}
                }
            )

            phase_name = phase.__name__
            logger.debug(f"Running phase: {phase_name}\n")
            phase_params = inspect.signature(phase).parameters

            # Prepare the arguments for the current phase
            args = [analysis_data]
            if idx == 0:
                result = phase(*args)
            elif analysis_type == 'content' and phase_name == 'phase5':
                result = phase(analysis_data, responses['phase1'])
            elif analysis_type == 'thematic' and phase_name == 'phase5':
                result = phase(analysis_data, responses['phase4'])
            elif analysis_type == 'thematic' and phase_name == 'phase6':
                result = phase(analysis_data, responses['phase2'])
            elif analysis_type == 'thematic' and phase_name == 'phase7':
                result = phase(analysis_data, responses['phase1'])
            else:
                result = phase(*args)

            if isinstance(result, tuple):
                response_json, formatted_prompt = result[0], result[1]
                if len(result) == 4:  # Last phase with different return signature
                    response_json, formatted_prompt, analysis_status, deletion_results = result
            else:
                response_json = result

            if response_json is None:
                logger.error(f"Phase {phase_name} returned None\n")
                continue

            responses[phase_name] = response_json
            formatted_prompts.append(formatted_prompt)

            tables = generate_tables_from_response(response_json)
            prompt_table_pairs.append({'prompt': formatted_prompt, 'tables': tables})

            # Check if the key 'table_format_visualization' is present in the response JSON
            if "table_format_visualization" in response_json:
                logger.debug(f"Key 'table_format_visualization' found in response_json\n")
                try:
                    logger.debug(f"Flowchart recognized\n")
                    # Generate a combined flowchart based on the response JSON
                    flowchart = create_combined_flowchart(response_json)
                    if flowchart:
                        save_flowchart_as_png(flowchart)
                        if 'DYNO' in os.environ:  # Check if running on Heroku
                            flowchart_path = f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/media/flowcharts/flowchart.png"
                        else:  # Local environment
                            flowchart_path = f"{settings.MEDIA_URL}flowcharts/flowchart.png"
                        logger.debug(f"Flowchart path updated: {flowchart_path}")
                except json.JSONDecodeError as e:
                    logger.error(f"Failed to parse response JSON: {e}\n")
                except Exception as e:
                    logger.error(f"Error generating flowchart: {e}\n")
            else:
                logger.debug(f"Key 'table_format_visualization' not found in response_json\n")

            if flowchart_path:
                logger.debug(f"Flowchart path updated: {flowchart_path}\n")

        logger.debug(f"Deletion results: {deletion_results}\n")

        analysis_result = {
            'prompt_table_pairs': prompt_table_pairs,
            'flowchart_path': flowchart_path,
            'analysis_status': analysis_status,
            'deletion_results': deletion_results
        }

        # Directly update the session within this function
        request = analysis_data.get('request')
        if request:
            session = request.session
            session['prompt_table_pairs'] = analysis_result.get('prompt_table_pairs', [])
            session['flowchart_path'] = analysis_result.get('flowchart_path', '')
            session['analysis_status'] = analysis_result.get('analysis_status', '')
            session['deletion_results'] = analysis_result.get('deletion_results', '')
            session.save()

        # Send the results to the WebSocket consumer
        await channel_layer.group_send(
            "analysis_group",
            {
                "type": "send_analysis_result",
                "content": analysis_result,
            }
        )

        return analysis_result
    logger.debug("Invalid analysis type\n")
    return {}

# Update the session data with analysis results via POST request.
@csrf_exempt
def update_session(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        request.session['prompt_table_pairs'] = data.get('prompt_table_pairs', [])
        request.session['flowchart_path'] = data.get('flowchart_path', '')
        request.session['analysis_status'] = data.get('analysis_status', '')
        request.session['deletion_results'] = data.get('deletion_results', '')

        return JsonResponse({'status': 'Session updated successfully'})
    return JsonResponse({'error': 'Invalid request'}, status=400)

# Render the dashboard view, handling setup and analysis initiation.
@login_required
def dashboard(request):
    # If the request method is GET, clear the session data
    if request.method == 'GET':
        clear_session_data(request)

    # Initialize the setup form, handling both POST data and file uploads
    setup_form = SetupForm(request.POST or None, request.FILES or None)

    # Retrieve parameters from the POST data or the session
    analysis_type = request.POST.get('analysis_type', request.session.get('analysis_type', ''))
    user_prompt = request.POST.get('user_prompt', request.session.get('user_prompt', ''))
    file_name = request.session.get('file_name', '')
    uploaded_file_name = request.session.get('uploaded_file_name', '')
    file_id = request.session.get('file_id')
    model_choice = request.session.get('model_choice')

    if analysis_type:
        request.session['analysis_type'] = analysis_type

    # Prepare the context dictionary with data to be passed to the template
    context = {
        'setup_form': setup_form,
        'uploaded_file_name': uploaded_file_name,
        'version': __version__,
        'file_id': file_id,
        'model_choice': model_choice,
        'analysis_status': request.session.get('analysis_status', ''),
        'deletion_results': request.session.get('deletion_results', ''),
        'analysis_type': request.session.get('analysis_type', ''),
        'user_prompt': user_prompt,
        'file_name': file_name,
        'tables': request.session.get('tables', []),
        'prompt_table_pairs': request.session.get('prompt_table_pairs', []),  # Pairs of prompts and their corresponding tables
        'prompt_table_pairs_json': json.dumps(request.session.get('prompt_table_pairs', []), cls=DjangoJSONEncoder),
        'flowchart_path': request.session.get('flowchart_path', '')
    }

    # If the request method is POST, handle the form submission
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'analyze':
            logger.debug("Analyze action triggered\n")
            setup_success = handle_setup(request, setup_form)
            if setup_success:
                # Prepare the data needed for the analysis task
                analysis_data = {
                    'file_id': file_id,
                    'model_choice': model_choice,
                    'analysis_type': analysis_type,
                    'assistant_id': request.session.get('assistant_id'),
                    'thread_id': request.session.get('thread_id'),
                    'vector_store_id': request.session.get('vector_store_id'),
                    'file_name': request.session.get('file_name'),
                    'user_prompt': request.session.get('user_prompt')
                }
                logger.debug(f"Dispatching analysis task: {analysis_data}\n")
                # Get the channel layer for asynchronous communication and send the analysis task to the appropriate channel
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.send)("analysis_channel", {
                    "type": "run_analysis",
                    "analysis_data": analysis_data
                })
                request.session.save()  # Save the session data
            else:
                context['analysis_status'] = request.session.get('analysis_status', '')

    # Render the dashboard template with the context data
    return render(request, 'qda_gpt/dashboard.html', context)
